import argparse
import os
import re
import shutil
from pathlib import Path

import xlrd
import xlutils.copy
import xlwt

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def a1_to_rc(a1: str) -> tuple[int, int]:
    m = CELL_RE.match(a1.strip())
    if not m:
        raise ValueError(f"Bad cell address: {a1}")
    col_letters = m.group(1).upper()
    row_num = int(m.group(2))
    col_idx = 0
    for ch in col_letters:
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    col_idx -= 1  # 0-based
    row_idx = row_num - 1  # 0-based
    return row_idx, col_idx


def coerce_value(s: str):
    s = s.strip()
    if s.lower() in ("true", "false"):
        return s.lower() == "true"
    try:
        if "." in s.replace(",", "."):
            return float(s.replace(",", "."))
        return int(s)
    except ValueError:
        return s


def edit_xlsx(src: Path, dst: Path, edits: list[tuple[int, int, object]], sheet_index: int):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for .xlsx edits")
    shutil.copy2(src, dst)
    wb = load_workbook(dst)
    ws = wb.worksheets[sheet_index]
    for r, c, val in edits:
        ws.cell(r + 1, c + 1).value = val
    wb.save(dst)


def edit_xls(src: Path, dst: Path, edits: list[tuple[int, int, object]], sheet_index: int):
    shutil.copy2(src, dst)
    rd_book = xlrd.open_workbook(dst, formatting_info=True)
    wb_copy = xlutils.copy.copy(rd_book)
    ws_copy = wb_copy.get_sheet(sheet_index)
    rd_sheet = rd_book.sheet_by_index(sheet_index)

    def get_num_fmt(r: int, c: int) -> str:
        try:
            xf_index = rd_sheet.cell_xf_index(r, c)
            xf = rd_book.xf_list[xf_index]
            fmt = rd_book.format_map.get(xf.format_key)
            return getattr(fmt, 'format_str', 'general')
        except Exception:
            return 'general'

    for r, c, val in edits:
        num_fmt = get_num_fmt(r, c)
        style = xlwt.easyxf(num_format_str=num_fmt)
        ws_copy.write(r, c, val, style)
    wb_copy.save(dst)


def parse_edits(pairs: list[str]) -> list[tuple[int, int, object]]:
    result = []
    for pair in pairs:
        if "=" not in pair:
            raise ValueError(f"Bad --set value (expected A1=value): {pair}")
        cell, value = pair.split("=", 1)
        r, c = a1_to_rc(cell)
        result.append((r, c, coerce_value(value)))
    return result


def main():
    parser = argparse.ArgumentParser(description="Edit specific Excel cells in a copied file")
    parser.add_argument("source", help="Source Excel file (.xls/.xlsx)")
    parser.add_argument("dest", help="Destination Excel file path")
    parser.add_argument("--sheet", type=int, default=0, help="Sheet index (0-based)")
    parser.add_argument("--set", dest="sets", action="append", default=[], help="Edit in form A1=value. Can be repeated")
    args = parser.parse_args()

    src = Path(args.source)
    dst = Path(args.dest)
    if not src.exists():
        raise SystemExit(f"Source not found: {src}")
    if not args.sets:
        raise SystemExit("Provide at least one --set A1=value")

    edits = parse_edits(args.sets)

    ext = src.suffix.lower()
    if ext == ".xlsx":
        edit_xlsx(src, dst, edits, args.sheet)
    elif ext == ".xls":
        edit_xls(src, dst, edits, args.sheet)
    else:
        raise SystemExit("Unsupported file type. Use .xls or .xlsx")

    print(f"Edited {len(edits)} cell(s) -> {dst}")


if __name__ == "__main__":
    main()









