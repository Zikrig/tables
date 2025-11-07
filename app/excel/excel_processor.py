"""Основной модуль для работы с Excel файлами и генерации заказов"""
import os
import errno
from typing import Dict, List
import shutil
import zipfile
import tempfile
import xml.etree.ElementTree as ET
ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

 

class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.sheet_index = 0
        self.is_xlsx = file_path.lower().endswith('.xlsx')

    def read_file(self):
        if not (self.is_xlsx and OPENPYXL_AVAILABLE):
            raise RuntimeError("Поддерживается только формат .xlsx")
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.worksheet = self.workbook.worksheets[self.sheet_index]

    def get_cell_value(self, row: int, col: int):
        try:
            cell = self.worksheet.cell(row + 1, col + 1)
            return cell.value
        except (IndexError, AttributeError):
            return None

    def get_all_data(self) -> List[List]:
        data = []
        for row in self.worksheet.iter_rows(values_only=True):
            row_data = list(row)
            data.append(row_data)
        return data

    def write_file(self, output_path: str, data: List[List], quantity_col: int,
                   quantities: Dict[str, float], article_col: int = 0, start_row: int = 2,
                   price_col: int = None, sum_col: int = None, template_file: str = None,
                   total_row: int = None, total_count_enabled: bool = True):
        source_file = self.file_path
        is_xlsx_source = source_file.lower().endswith('.xlsx')
        base_name = os.path.splitext(os.path.basename(output_path))[0]

        if not (is_xlsx_source and OPENPYXL_AVAILABLE):
            raise RuntimeError("Поддерживается только формат .xlsx")
        if is_xlsx_source and OPENPYXL_AVAILABLE:
            # 1) Точная копия исходного файла
            shutil.copy2(source_file, output_path)
            # 2) Считаем список обновлений ячеек (A1 -> значение)
            updates: Dict[str, float] = {}
            def idx_to_col(col_index_zero_based: int) -> str:
                n = col_index_zero_based + 1
                s = ""
                while n > 0:
                    n, r = divmod(n - 1, 26)
                    s = chr(65 + r) + s
                return s
            def rc_to_a1(r0: int, c0: int) -> str:
                return f"{idx_to_col(c0)}{r0 + 1}"
            total_quantity = 0.0
            total_sum = 0.0
            for row_idx, row_data in enumerate(data):
                if row_idx < start_row:
                    continue
                article = None
                if row_data and len(row_data) > article_col and row_data[article_col]:
                    article = _normalize_article(row_data[article_col])
                # Use normalized keys for lookup
                norm_quantities: Dict[str, float] = { _normalize_article(k): v for k, v in quantities.items() if _normalize_article(k) }
                if article and article in norm_quantities:
                    qty = float(norm_quantities[article])
                    if qty > 0:
                        updates[rc_to_a1(row_idx, quantity_col)] = qty
                        total_quantity += qty
                        if sum_col is not None and price_col is not None:
                            price = 0.0
                            if price_col < len(row_data) and row_data[price_col]:
                                try:
                                    price = float(row_data[price_col])
                                except (ValueError, TypeError):
                                    price = 0.0
                            if price > 0:
                                row_sum = price * qty
                                updates[rc_to_a1(row_idx, sum_col)] = row_sum
                                total_sum += row_sum
            if total_row is not None and total_row >= 0:
                if total_count_enabled and quantity_col is not None:
                    updates[rc_to_a1(total_row, quantity_col)] = total_quantity
                if sum_col is not None:
                    updates[rc_to_a1(total_row, sum_col)] = total_sum
            # 3) Обновляем ячейки через openpyxl (как в test/edit_cells.py)
            wb = load_workbook(output_path)
            ws = wb.worksheets[self.sheet_index]
            for a1_ref, number in updates.items():
                ws[a1_ref].value = number
            wb.save(output_path)
            # 4) Готово

    def close(self) -> None:
        """Совместимость: закрытие ресурсов (ничего не делает)."""
        return


def convert_to_xlsx(input_path: str, output_path: str) -> str:
    raise RuntimeError("Поддержка .xls отключена. Загрузите файл в формате .xlsx.")

    # unreachable

    # end

    # noqa

def _ns() -> Dict[str, str]:
    return {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

def _to_plain_string(value) -> str:
    # Extract plain text from rich text objects if present; fall back to str(value)
    try:
        # openpyxl rich text may have 'runs' iterable of text blocks
        if hasattr(value, 'runs'):
            try:
                return "".join(getattr(r, 'text', str(r)) for r in value.runs)
            except Exception:
                pass
        # Many objects expose .text
        if hasattr(value, 'text'):
            t = getattr(value, 'text')
            if t is not None:
                return str(t)
    except Exception:
        pass
    return str(value)

def _normalize_article(value) -> str:
    if value is None:
        return ""
    s = _to_plain_string(value)
    # normalize spaces (incl. NBSP/thin NBSP), strip, then remove all whitespace
    s = s.replace('\u00A0', ' ').replace('\u202F', ' ')
    s = s.strip()
    if not s:
        return ""
    # Remove all whitespace to avoid formatting mismatches and unify case
    s = "".join(ch for ch in s if not ch.isspace())
    return s

def normalize_article(value) -> str:
    return _normalize_article(value)

def _coerce_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    s = str(value)
    # normalize non-breaking spaces and thin spaces
    s = s.replace('\u00A0', ' ').replace('\u202F', ' ')
    s = s.strip()
    if not s:
        return None
    # If both ',' and '.' present, the last one is decimal; remove others
    last_comma = s.rfind(',')
    last_dot = s.rfind('.')
    try:
        if last_comma == -1 and last_dot == -1:
            # No explicit decimal separator: remove spaces and parse
            s2 = s.replace(' ', '')
            return float(s2)
        if last_comma > last_dot:
            # Comma as decimal
            s2 = s.replace('.', '')
            s2 = s2.replace(' ', '')
            s2 = s2.replace(',', '.')
            return float(s2)
        else:
            # Dot as decimal
            s2 = s.replace(',', '')
            s2 = s2.replace(' ', '')
            return float(s2)
    except ValueError:
        return None

def collect_article_quantities_xlsx(file_path: str, sheet_index: int, article_col_letter: str, quantity_col_letter: str, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
        raise RuntimeError("Поддерживается только формат .xlsx")
    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb.worksheets[sheet_index]
        result: Dict[str, float] = {}
        max_row = min(ws.max_row or 0, max_rows)
        a_col = column_index_from_string(article_col_letter)
        q_col = column_index_from_string(quantity_col_letter)
        # start_row is 1-based (2 means skip header)
        start_r = max(1, start_row)
        for r in range(start_r, max_row + 1):
            art_raw = ws.cell(r, a_col).value
            qty_raw = ws.cell(r, q_col).value
            article = _normalize_article(art_raw)
            if not article:
                continue
            qty = _coerce_float(qty_raw)
            if qty is None:
                continue
            if qty == 0:
                continue
            prev = result.get(article, 0.0)
            result[article] = prev + float(qty)
        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass

def get_warehouse_articles(file_path: str, sheet_index: int = 0, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    return collect_article_quantities_xlsx(file_path, sheet_index, 'A', 'E', max_rows, start_row)

def get_preorder_articles(file_path: str, sheet_index: int = 0, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    return collect_article_quantities_xlsx(file_path, sheet_index, 'C', 'E', max_rows, start_row)

def collect_article_quantities_xlsx_all_sheets(file_path: str, article_col_letter: str, quantity_col_letter: str, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
        raise RuntimeError("Поддерживается только формат .xlsx")
    wb = load_workbook(file_path, data_only=True)
    try:
        result: Dict[str, float] = {}
        a_col = column_index_from_string(article_col_letter)
        q_col = column_index_from_string(quantity_col_letter)
        start_r = max(1, start_row)
        for ws in wb.worksheets:
            max_row = min(ws.max_row or 0, max_rows)
            for r in range(start_r, max_row + 1):
                art_raw = ws.cell(r, a_col).value
                qty_raw = ws.cell(r, q_col).value
                article = _normalize_article(art_raw)
                if not article:
                    continue
                qty = _coerce_float(qty_raw)
                if qty is None or qty == 0:
                    continue
                prev = result.get(article, 0.0)
                result[article] = prev + float(qty)
        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass

def get_warehouse_articles_all_sheets(file_path: str, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    return collect_article_quantities_xlsx_all_sheets(file_path, 'A', 'E', max_rows, start_row)

def get_preorder_articles_all_sheets(file_path: str, max_rows: int = 1000, start_row: int = 2) -> Dict[str, float]:
    return collect_article_quantities_xlsx_all_sheets(file_path, 'C', 'E', max_rows, start_row)

def _print_articles_with_sheet_all_sheets(file_path: str, article_col_letter: str, quantity_col_letter: str, max_rows: int = 1000, start_row: int = 2) -> None:
    if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
        raise RuntimeError("Поддерживается только формат .xlsx")
    wb = load_workbook(file_path, data_only=True)
    try:
        a_col = column_index_from_string(article_col_letter)
        q_col = column_index_from_string(quantity_col_letter)
        start_r = max(1, start_row)
        for ws in wb.worksheets:
            max_row = min(ws.max_row or 0, max_rows)
            for r in range(start_r, max_row + 1):
                art_raw = ws.cell(r, a_col).value
                qty_raw = ws.cell(r, q_col).value
                article = _normalize_article(art_raw)
                if not article:
                    continue
                qty = _coerce_float(qty_raw)
                if qty is None or qty == 0:
                    continue
                print(f"{article}\t{ws.title}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

def print_warehouse_articles_with_sheets(file_path: str, max_rows: int = 1000, start_row: int = 2) -> None:
    _print_articles_with_sheet_all_sheets(file_path, 'A', 'E', max_rows, start_row)

def print_preorder_articles_with_sheets(file_path: str, max_rows: int = 1000, start_row: int = 2) -> None:
    _print_articles_with_sheet_all_sheets(file_path, 'C', 'E', max_rows, start_row)

def _col_row_from_a1(a1: str) -> (str, int):
    i = 0
    while i < len(a1) and a1[i].isalpha():
        i += 1
    col = a1[:i]
    row = int(a1[i:]) if i < len(a1) else 0
    return col, row

def _find_row(sheet_data_el, row_number: int, ns: Dict[str, str]):
    for row in sheet_data_el.findall('a:row', ns):
        r = row.get('r')
        if r and int(r) == row_number:
            return row
    return None

def _find_cell(row_el, a1_ref: str, ns: Dict[str, str]):
    for c in row_el.findall('a:c', ns):
        if c.get('r') == a1_ref:
            return c
    return None

def _set_cell_number(cell_el, value: float, ns: Dict[str, str]):
    # Меняем только существующие числовые ячейки: t отсутствует или t == 'n'
    t = cell_el.get('t')
    if t is not None and t != 'n':
        return
    v = cell_el.find('a:v', ns)
    if v is None:
        return
    v.text = str(value)

def _patch_xml_bytes(sheet_xml_bytes: bytes, updates: Dict[str, float]) -> bytes:
    ns = _ns()
    root = ET.fromstring(sheet_xml_bytes)
    sheet_data = root.find('a:sheetData', ns)
    if sheet_data is None:
        return sheet_xml_bytes
    # Группируем по строкам
    row_to_cells: Dict[int, Dict[str, float]] = {}
    for a1, val in updates.items():
        col, row = _col_row_from_a1(a1)
        if row <= 0:
            continue
        row_to_cells.setdefault(row, {})[a1] = val
    for row_num, cell_map in row_to_cells.items():
        row_el = _find_row(sheet_data, row_num, ns)
        if row_el is None:
            # Не создаём новые строки, чтобы не трогать стили
            continue
        for a1_ref, number in cell_map.items():
            cell_el = _find_cell(row_el, a1_ref, ns)
            if cell_el is None:
                # Не создаём новые ячейки, чтобы не ломать стиль/форматирование
                continue
            _set_cell_number(cell_el, number, ns)
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)

def _sheet_xml_name(sheet_index_zero_based: int) -> str:
    # Предполагаем sheetN.xml, где N = index+1
    return f'xl/worksheets/sheet{sheet_index_zero_based + 1}.xml'

def _rewrite_zip_with_replacement(src_path: str, dst_path: str, member_name: str, new_bytes: bytes) -> None:
    with zipfile.ZipFile(src_path, 'r') as zin, zipfile.ZipFile(dst_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == member_name:
                zout.writestr(item, new_bytes)
            else:
                zout.writestr(item, data)

def _atomic_replace(path: str, tmp_path: str) -> None:
    try:
        os.replace(tmp_path, path)
    except OSError as e:
        if getattr(e, "errno", None) == errno.EXDEV:
            # Кросс-девайс: используем безопасный move
            shutil.move(tmp_path, path)
        else:
            raise

def _read_zip_member(path: str, member_name: str) -> bytes:
    with zipfile.ZipFile(path, 'r') as zf:
        return zf.read(member_name)

def _write_zip_member_inplace(path: str, member_name: str, content: bytes) -> None:
    # Создаём временный архив и заменяем (во временном файле в той же директории, чтобы избежать EXDEV)
    tmp_dir = os.path.dirname(os.path.abspath(path)) or "."
    with tempfile.NamedTemporaryFile(delete=False, dir=tmp_dir) as tmp:
        tmp_path = tmp.name
    try:
        _rewrite_zip_with_replacement(path, tmp_path, member_name, content)
        _atomic_replace(path, tmp_path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
        except Exception:
            pass

def _cell_exists_in_sheet_xml(sheet_xml_bytes: bytes, a1_ref: str) -> bool:
    ns = _ns()
    root = ET.fromstring(sheet_xml_bytes)
    sheet_data = root.find('a:sheetData', ns)
    if sheet_data is None:
        return False
    col, row = _col_row_from_a1(a1_ref)
    row_el = _find_row(sheet_data, row, ns)
    if row_el is None:
        return False
    return _find_cell(row_el, a1_ref, ns) is not None

def _filter_updates_to_existing_cells(sheet_xml_bytes: bytes, updates: Dict[str, float]) -> Dict[str, float]:
    filtered: Dict[str, float] = {}
    for a1, val in updates.items():
        if _cell_exists_in_sheet_xml(sheet_xml_bytes, a1):
            filtered[a1] = val
    return filtered

def ExcelProcessor__patch_sheet_xml(self, xlsx_path: str, sheet_index_zero_based: int, updates: Dict[str, float]) -> None:
    sheet_name = _sheet_xml_name(sheet_index_zero_based)
    original = _read_zip_member(xlsx_path, sheet_name)
    # Пишем только в существующие ячейки, чтобы не терять стиль пустых
    updates2 = _filter_updates_to_existing_cells(original, updates)
    if not updates2:
        return
    patched = _patch_xml_bytes(original, updates2)
    _write_zip_member_inplace(xlsx_path, sheet_name, patched)

# Привязываем как метод класса
ExcelProcessor._patch_sheet_xml = ExcelProcessor__patch_sheet_xml


