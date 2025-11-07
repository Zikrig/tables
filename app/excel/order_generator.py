
"""Модуль для сопоставления товаров и генерации заказов"""
from typing import Dict, Optional
import logging
from app.excel.excel_processor import ExcelProcessor, normalize_article as _norm_article
from app.excel.excel_processor import _coerce_float as _coerce_qty
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class OrderGenerator:
    def __init__(self, price_config: Dict):
        self.price_config = price_config
        self.logger = logging.getLogger(__name__)
        self.last_diagnostics: Dict[str, object] = {}

    def read_price_list(self, file_path: str) -> Dict[str, Dict]:
        processor = ExcelProcessor(file_path)
        processor.read_file()
        data = processor.get_all_data()
        price_items = {}
        start_row = self.price_config.get('start_row', 2)
        article_col = self.price_config.get('article_col', 0)
        for row_idx in range(start_row, len(data)):
            row = data[row_idx]
            if article_col < len(row) and row[article_col]:
                article = str(row[article_col]).strip()
                if article:
                    price_items[article] = {'row': row_idx, 'data': row}
        processor.close()
        return price_items

    def read_warehouse_order(self, file_path: str, config: Dict) -> Dict[str, float]:
        if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
            # Фолбек на старый путь через ExcelProcessor (первый лист)
            processor = ExcelProcessor(file_path)
            processor.read_file()
            data = processor.get_all_data()
            processor.close()
            quantities: Dict[str, float] = {}
            # Skip exactly one header row
            start_row = 1
            article_col = self.price_config.get('article_col', 0)
            quantity_col = self.price_config.get('quantity_col', 9)
            for row_idx in range(start_row, len(data)):
                row = data[row_idx]
                if article_col < len(row) and row[article_col]:
                    article = str(row[article_col]).strip()
                    if not article:
                        continue
                    qty = 0.0
                    if quantity_col < len(row) and row[quantity_col]:
                        try:
                            qty = float(row[quantity_col])
                        except (ValueError, TypeError):
                            qty = 0.0
                    if qty > 0:
                        quantities[article] = quantities.get(article, 0.0) + qty
            return quantities
        # Новый путь: обходим все листы, суммируем, ограничиваем 1000 строк
        wb = load_workbook(file_path, data_only=True)
        try:
            quantities: Dict[str, float] = {}
            # Skip exactly one header row (1-based in Excel)
            start_row = 1
            article_col = self.price_config.get('article_col', 0)
            quantity_col = self.price_config.get('quantity_col', 9)
            max_rows = 1000
            total_rows_seen = 0
            total_articles_seen = 0
            total_valid_qty_rows = 0
            for ws in wb.worksheets:
                max_row = min(ws.max_row or 0, max_rows)
                for r in range(2, max_row + 1):  # start from row 2 (skip only header)
                    art_val = ws.cell(r, article_col + 1).value
                    qty_val = ws.cell(r, quantity_col + 1).value
                    article = _norm_article(art_val)
                    if not article:
                        total_rows_seen += 1
                        continue
                    total_rows_seen += 1
                    total_articles_seen += 1
                    qty = _coerce_qty(qty_val) or 0.0
                    if qty > 0:
                        total_valid_qty_rows += 1
                    if qty > 0:
                        quantities[article] = quantities.get(article, 0.0) + qty
            self.last_diagnostics['warehouse'] = {
                'article_col_index': article_col,
                'quantity_col_index': quantity_col,
                'rows_seen': total_rows_seen,
                'articles_seen': total_articles_seen,
                'valid_qty_rows': total_valid_qty_rows,
                'total_items_found': len(quantities)
            }
            return quantities
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def read_preorders(self, file_path: str, config: Dict) -> Dict[str, float]:
        if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
            # Фолбек: только первый лист
            processor = ExcelProcessor(file_path)
            processor.read_file()
            data = processor.get_all_data()
            processor.close()
            quantities: Dict[str, float] = {}
            # Skip exactly one header row
            start_row = 1
            article_col = config.get('article_col', 2)
            article_col2 = config.get('article_col2', 5)
            quantity_col = config.get('quantity_col', 4)
            for row_idx in range(start_row, len(data)):
                row = data[row_idx]
                if not row:
                    continue
                article = None
                if article_col2 < len(row) and row[article_col2]:
                    article = str(row[article_col2]).strip()
                if not article and article_col < len(row) and row[article_col]:
                    article = str(row[article_col]).strip()
                if not article:
                    continue
                qty = 0.0
                if quantity_col < len(row) and row[quantity_col]:
                    try:
                        qty = float(row[quantity_col])
                    except (ValueError, TypeError):
                        qty = 0.0
                if qty > 0:
                    quantities[article] = quantities.get(article, 0.0) + qty
            return quantities
        # Новый путь: все листы, суммирование, ограничение 1000 строк
        wb = load_workbook(file_path, data_only=True)
        try:
            quantities: Dict[str, float] = {}
            # Skip exactly one header row (1-based in Excel)
            start_row = 1
            article_col = config.get('article_col', 2)
            article_col2 = config.get('article_col2', 5)
            quantity_col = config.get('quantity_col', 4)
            max_rows = 1000
            total_rows_seen = 0
            total_articles_seen = 0
            total_valid_qty_rows = 0
            for ws in wb.worksheets:
                max_row = min(ws.max_row or 0, max_rows)
                for r in range(2, max_row + 1):
                    art1 = ws.cell(r, article_col + 1).value
                    art2 = ws.cell(r, article_col2 + 1).value
                    article = _norm_article(art2) or _norm_article(art1)
                    if not article:
                        total_rows_seen += 1
                        continue
                    total_rows_seen += 1
                    total_articles_seen += 1
                    qty_val = ws.cell(r, quantity_col + 1).value
                    qty = _coerce_qty(qty_val) or 0.0
                    if qty > 0:
                        total_valid_qty_rows += 1
                    if qty > 0:
                        quantities[article] = quantities.get(article, 0.0) + qty
            self.last_diagnostics['preorders'] = {
                'article_col_index': article_col,
                'article_col2_index': article_col2,
                'quantity_col_index': quantity_col,
                'rows_seen': total_rows_seen,
                'articles_seen': total_articles_seen,
                'valid_qty_rows': total_valid_qty_rows,
                'total_items_found': len(quantities)
            }
            return quantities
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def preview_warehouse(self, file_path: str, article_col: int, quantity_col: int, rows: int = 10) -> str:
        if not (file_path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE):
            return "(preview unavailable: not .xlsx)"
        wb = load_workbook(file_path, data_only=True)
        try:
            lines = []
            for ws in wb.worksheets:
                max_row = min(ws.max_row or 0, 1 + rows)
                lines.append(f"Лист: {ws.title}")
                for r in range(2, max_row + 1):
                    a_raw = ws.cell(r, article_col + 1).value
                    q_raw = ws.cell(r, quantity_col + 1).value
                    a_norm = _norm_article(a_raw)
                    q_num = _coerce_qty(q_raw)
                    lines.append(
                        f"r{r}: col{article_col + 1}='" + str(a_raw) + "' -> '" + a_norm + "', "
                        f"col{quantity_col + 1}='" + str(q_raw) + "' -> " + str(q_num)
                    )
                # только первый лист, чтобы не засорять
                break
            return "\n".join(lines)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def generate_order(self, price_file: str, warehouse_file: str, preorders_file: str,
                      output_file: str, warehouse_config: Dict, preorders_config: Dict,
                      price_template: Optional[str] = None):
        price_items = self.read_price_list(price_file)
        warehouse_quantities = self.read_warehouse_order(warehouse_file, warehouse_config)
        preorder_quantities = self.read_preorders(preorders_file, preorders_config)
        final_quantities: Dict[str, float] = {}
        all_articles = set(list(warehouse_quantities.keys()) + list(preorder_quantities.keys()))
        for article in all_articles:
            warehouse_qty = warehouse_quantities.get(article, 0)
            preorder_qty = preorder_quantities.get(article, 0)
            if warehouse_qty == 0:
                final_qty = preorder_qty
            elif preorder_qty == 0:
                final_qty = warehouse_qty
            else:
                final_qty = warehouse_qty + preorder_qty
            if final_qty > 0:
                final_quantities[article] = final_qty
        processor = ExcelProcessor(price_file)
        processor.read_file()
        data = processor.get_all_data()
        quantity_col = self.price_config.get('quantity_col', 9)
        article_col = self.price_config.get('article_col', 0)
        price_col = self.price_config.get('price_col')
        sum_col = self.price_config.get('sum_col')
        start_row = self.price_config.get('start_row', 2)
        total_row = None
        total_count_enabled = False
        # Для сохранения форматирования используем исходный шаблон, если он есть
        template = price_template or price_file
        processor.write_file(output_file, data, quantity_col, final_quantities,
                             article_col, start_row, price_col, sum_col, template, total_row, total_count_enabled)
        processor.close()
        return final_quantities


